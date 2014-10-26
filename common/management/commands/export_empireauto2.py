# -*- coding=utf-8 -*-

from django.db import connection
from django.core.management.base import BaseCommand


class Command(BaseCommand):
    option_list = BaseCommand.option_list + (
        
    )

    def handle(self, *args, **options):

        import openpyxl

        sheet = 1

        wb = openpyxl.Workbook(optimized_write=True)
        ws = wb.create_sheet(sheet)

        ws.append((u'Брэнд', u'Артикул', u'Наименование', u'Аббревиатура', u'List', u'Вход', u'Партия', u'Наличие'))

        cur = connection.cursor()

        sql = """
        SELECT DISTINCT ON (maker, refoenumber, itemnumber) itemnumber, yourprice, description
        FROM common_bot2
        ORDER BY itemnumber
        """

        cur.execute(sql)

        i = 1
        for item in cur:
            itemnumber, yourprice, description = item

            if yourprice:
                yourprice = float(yourprice)
                invoice = yourprice * 1.05
                ws.append((
                    'Taiwan', itemnumber, '', description,
                    '%.2f' % yourprice, '%.2f' % invoice, 1, 0))
                i += 1

            if i == 65000:
                sheet += 1
                ws = wb.create_sheet(sheet)
                ws.append((u'Брэнд', u'Артикул', u'Наименование', u'Аббревиатура', u'List', u'Вход', u'Партия', u'Наличие'))
                i = 1

            print i

        wb.save('export2.xlsx')
