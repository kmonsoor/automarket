# -*- coding=utf-8 -*-

import json



from django.db import connection
from django.core.management.base import BaseCommand


MAP = {
    'ACURA': 'Acura',
    'AUDI': 'Audi',
    'BMW': 'BMW',
    'BUICK':'Gm',
    'CADILLAC':'Gm',
    'CHEVROLET-GMC':'Gm',
    'CHRYSLER-DODGE-PLYM':'Mopar',
    'DAEWOO': 'Daewoo',
    'EAGLE': 'Mopar',
    'FIAT': 'Fiat',
    'FORD-MERCURY':'Ford',
    'GEO':'Gm',
    'HUMMER': 'Gm',
    'HONDA':'Honda',
    'HYUNDAI':'Hyundai',
    'INFINITI': 'Infiniti',
    'ISUZU':'Isuzu',
    'JAGUAR': 'Jaguar',
    'JEEP':'Mopar',
    'KIA': 'Kia',
    'LAND ROVER': 'Land Rover',
    'LEXUS': 'Lexus',
    'LINCOLN':'Ford',
    'MAZDA':'Mazda',
    'MINI': 'Bmw',
    'MERCEDES BENZ':'Mercedes-benz',
    'MITSUBISHI':'Mitsubishi',
    'NISSAN':'Nissan',
    'OLDSMOBILE':'Gm',
    'PONTIAC':'Gm',
    'PORSCHE': 'Porsche',
    'SAAB':'Gm',
    'SATURN': 'Gm',
    'SCION': 'Toyota',
    'SUBARU':'Subaru',
    'SUZUKI':'Suzuki',
    'TOYOTA':'Toyota',
    'VOLKSWAGEN':'Volkswagen',
    'VOLVO':'Volvo',
}


class Command(BaseCommand):
    option_list = BaseCommand.option_list + (
        
    )

    def handle(self, *args, **options):

        import openpyxl

        sheet = 1

        wb = openpyxl.Workbook(optimized_write=True)
        ws = wb.create_sheet(sheet)

        ws.append((u'Брэнд1', u'Артикул1', u'Брэнд2', u'Артикул2',))

        cur = connection.cursor()

        sql = """
        SELECT DISTINCT ON (maker, refoenumber, itemnumber) maker, refoenumber, itemnumber, assembly
        FROM common_bot2
        ORDER BY maker
        """

        cur.execute(sql)

        i = 1
        for item in cur:
            maker, refoenumber, itemnumber, assemblies = item

            obrand = MAP[maker.upper()]
            onumber = refoenumber
            fbrand = 'Taiwan'
            fnumber = itemnumber

            if onumber == 'Assembly':
                try:
                    assemblies = json.loads(assemblies)
                except:
                    continue
                for assembly in assemblies:
                    if assembly[0] == 'OEM #':
                        row = (obrand, assembly[1], fbrand, fnumber,)
                        if all(row):
                            ws.append(row)
                            i += 1
            else:
                row1 = (obrand, onumber, fbrand, fnumber,)
                if all(row1):
                    ws.append(row1)
                    i += 1
                row2 = (fbrand, fnumber, obrand, onumber,)
                if all(row2):
                    ws.append(row2)
                    i += 1

            if i == 65000:
                sheet += 1
                ws = wb.create_sheet(sheet)
                ws.append((u'Брэнд1', u'Артикул1', u'Брэнд2', u'Артикул2',))
                i = 1

            print i

        wb.save('export1.xlsx')
